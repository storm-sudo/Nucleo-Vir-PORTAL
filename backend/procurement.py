# Procurement to Payment Automation Module
# procurement.py

from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Cookie, Query
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel, ConfigDict
from typing import Optional, List, Dict, Any
from datetime import datetime, timezone
from motor.motor_asyncio import AsyncIOMotorClient
import uuid
import os
import io
import json
import re
from pathlib import Path

# PDF and OCR imports
try:
    import pytesseract
    from PIL import Image
    from pdf2image import convert_from_bytes
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
    from reportlab.pdfgen import canvas
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# Router
procurement_router = APIRouter(prefix="/api", tags=["Procurement"])

# Configuration
CA_EMAIL = "nikita@nucleovir.com"
DIRECTORS = [
    "yogesh.ostwal@nucleovir.com",
    "sunil.k@nucleovir.com",
    "ayush@nucleovir.com"
]
DIRECTOR_PRIORITY = {
    "yogesh.ostwal@nucleovir.com": 1,
    "sunil.k@nucleovir.com": 2,
    "ayush@nucleovir.com": 3
}

# Approval thresholds (in INR)
THRESHOLDS = {
    "single_director": 50000,
    "single_director_3quotes": 200000,
    "all_directors": 1000000,
    "board_approval": 10000000
}

# Storage paths
STORAGE_BASE = "/app/storage"
FY_YEAR = "FY2026"

# Pydantic Models
class QuotationCreate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    quotation_no: str
    vendor_name: str
    category: str
    gst_pct: float = 18.0
    total_amount: float
    validity_date: str
    department: str
    description: Optional[str] = ""
    items: Optional[List[Dict]] = []

class POGenerate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    quotation_id: str
    delivery_address: Optional[str] = "NucleoVir Therapeutics Pvt. Ltd."
    payment_terms: Optional[str] = "Net 30"
    notes: Optional[str] = ""

class ApprovalDecision(BaseModel):
    model_config = ConfigDict(extra="ignore")
    decision: str  # "approved" or "rejected"
    comment: Optional[str] = ""
    approver_email: str

class GRNCreate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    po_id: str
    received_items: List[Dict]
    qc_status: str = "Passed"
    qc_notes: Optional[str] = ""
    asset_tags: Optional[List[str]] = []

class VoucherCreate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    grn_id: str
    invoice_number: str
    invoice_date: str
    invoice_amount: float
    tds_pct: Optional[float] = 0.0
    payment_mode: Optional[str] = "Bank Transfer"

# Helper functions
def get_fiscal_year():
    now = datetime.now()
    if now.month >= 4:
        return f"FY{now.year + 1}"
    return f"FY{now.year}"

def get_storage_path(category: str, vendor_name: str = None, month: str = None):
    fy = get_fiscal_year()
    base = Path(STORAGE_BASE)
    
    paths = {
        "quotations": base / "Finance" / "Purchase" / "Quotations" / fy,
        "po": base / "Finance" / "Purchase" / "PO" / fy / (vendor_name or "General"),
        "invoice": base / "Finance" / "Expense" / "Reimbursement" / "Invoice",
        "voucher": base / "Finance" / "PaymentProof" / "Voucher" / fy / (month or datetime.now().strftime("%B")),
        "grn": base / "Finance" / "Purchase" / "GRN" / fy,
    }
    
    path = paths.get(category, base / "Finance" / "Misc")
    path.mkdir(parents=True, exist_ok=True)
    return path

def generate_po_number(db_counter: int):
    fy = get_fiscal_year()
    return f"NVT/PO/{fy}/{str(db_counter).zfill(4)}"

def generate_grn_number(db_counter: int):
    fy = get_fiscal_year()
    return f"NVT/GRN/{fy}/{str(db_counter).zfill(4)}"

def generate_voucher_number(db_counter: int):
    fy = get_fiscal_year()
    return f"NVT/VCH/{fy}/{str(db_counter).zfill(4)}"

async def create_audit_log(db, entity_type: str, entity_id: str, action: str, performed_by: str, old_value: dict = None, new_value: dict = None):
    """Create an audit log entry"""
    audit_doc = {
        "audit_id": f"audit_{uuid.uuid4().hex[:12]}",
        "entity_type": entity_type,
        "entity_id": entity_id,
        "action": action,
        "performed_by": performed_by,
        "old_value": old_value,
        "new_value": new_value,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "ip_address": None,
        "retention_until": datetime(datetime.now().year + 7, 12, 31).isoformat()
    }
    await db.audit_logs.insert_one(audit_doc)
    return audit_doc

async def create_notification(db, user_email: str, title: str, message: str, notification_type: str, related_entity: dict = None):
    """Create in-app notification"""
    notif_doc = {
        "notification_id": f"notif_{uuid.uuid4().hex[:12]}",
        "user_email": user_email.lower(),
        "title": title,
        "message": message,
        "type": notification_type,
        "related_entity": related_entity,
        "read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notif_doc)
    return notif_doc

async def create_approval_entries(db, po_id: str, amount: float, created_by: str):
    """Create approval entries based on amount thresholds"""
    approvals = []
    
    if amount <= THRESHOLDS["single_director"]:
        # Single director approval - Yogesh first
        approval = {
            "approval_id": f"appr_{uuid.uuid4().hex[:12]}",
            "po_id": po_id,
            "approver_email": DIRECTORS[0],
            "approval_order": 1,
            "status": "pending",
            "required": True,
            "decision": None,
            "comment": None,
            "decided_at": None,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        approvals.append(approval)
        
    elif amount <= THRESHOLDS["single_director_3quotes"]:
        # Single director + 3 quotes required
        approval = {
            "approval_id": f"appr_{uuid.uuid4().hex[:12]}",
            "po_id": po_id,
            "approver_email": DIRECTORS[0],
            "approval_order": 1,
            "status": "pending",
            "required": True,
            "requires_3_quotes": True,
            "decision": None,
            "comment": None,
            "decided_at": None,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        approvals.append(approval)
        
    elif amount <= THRESHOLDS["all_directors"]:
        # All directors must approve in order
        for i, director in enumerate(DIRECTORS):
            approval = {
                "approval_id": f"appr_{uuid.uuid4().hex[:12]}",
                "po_id": po_id,
                "approver_email": director,
                "approval_order": i + 1,
                "status": "pending" if i == 0 else "waiting",
                "required": True,
                "decision": None,
                "comment": None,
                "decided_at": None,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            approvals.append(approval)
    else:
        # Board approval required
        for i, director in enumerate(DIRECTORS):
            approval = {
                "approval_id": f"appr_{uuid.uuid4().hex[:12]}",
                "po_id": po_id,
                "approver_email": director,
                "approval_order": i + 1,
                "status": "pending" if i == 0 else "waiting",
                "required": True,
                "requires_board_resolution": True,
                "decision": None,
                "comment": None,
                "decided_at": None,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            approvals.append(approval)
    
    # Insert all approvals
    if approvals:
        await db.approvals.insert_many(approvals)
        
        # Notify first approver
        first_approver = approvals[0]["approver_email"]
        await create_notification(
            db, 
            first_approver,
            "New PO Pending Approval",
            f"A purchase order for ₹{amount:,.2f} requires your approval.",
            "approval_pending",
            {"type": "po", "id": po_id}
        )
    
    return approvals

def ocr_extract_from_pdf(pdf_bytes: bytes) -> dict:
    """Extract text from PDF using OCR"""
    if not OCR_AVAILABLE:
        return {"error": "OCR not available", "raw_text": "", "confidence": 0}
    
    try:
        # Convert PDF to images
        images = convert_from_bytes(pdf_bytes, dpi=200)
        
        full_text = ""
        for img in images:
            text = pytesseract.image_to_string(img)
            full_text += text + "\n"
        
        # Try to extract structured data
        extracted = {
            "raw_text": full_text,
            "confidence": 75,  # Default confidence
            "fields": {}
        }
        
        # Extract quotation number
        quo_match = re.search(r'(?:Quotation|Quote|Ref)[\s#:No.]*([A-Z0-9/-]+)', full_text, re.IGNORECASE)
        if quo_match:
            extracted["fields"]["quotation_no"] = {"value": quo_match.group(1), "confidence": 80}
        
        # Extract vendor name (usually at top)
        lines = full_text.split('\n')
        for line in lines[:5]:
            if len(line.strip()) > 5 and not any(x in line.lower() for x in ['quotation', 'date', 'to:', 'from:']):
                extracted["fields"]["vendor_name"] = {"value": line.strip(), "confidence": 60}
                break
        
        # Extract amounts
        amount_matches = re.findall(r'(?:Rs\.?|INR|₹)\s*([\d,]+(?:\.\d{2})?)', full_text)
        if amount_matches:
            amounts = [float(a.replace(',', '')) for a in amount_matches]
            extracted["fields"]["total_amount"] = {"value": max(amounts), "confidence": 70}
        
        # Extract GST
        gst_match = re.search(r'GST\s*[@:]?\s*(\d+)%', full_text, re.IGNORECASE)
        if gst_match:
            extracted["fields"]["gst_pct"] = {"value": float(gst_match.group(1)), "confidence": 85}
        
        # Extract date
        date_match = re.search(r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})', full_text)
        if date_match:
            extracted["fields"]["validity_date"] = {"value": date_match.group(1), "confidence": 70}
        
        return extracted
        
    except Exception as e:
        return {"error": str(e), "raw_text": "", "confidence": 0}

def generate_po_pdf(po_data: dict) -> bytes:
    """Generate PO PDF"""
    if not PDF_AVAILABLE:
        return None
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=20)
    normal_style = styles['Normal']
    
    elements = []
    
    # Header
    elements.append(Paragraph("NucleoVir Therapeutics Pvt. Ltd.", title_style))
    elements.append(Paragraph("<b>PURCHASE ORDER</b>", ParagraphStyle('POTitle', parent=styles['Heading2'], alignment=1)))
    elements.append(Spacer(1, 20))
    
    # PO Details
    po_info = [
        ["PO Number:", po_data.get("po_number", "")],
        ["Date:", po_data.get("created_at", "")[:10]],
        ["Vendor:", po_data.get("vendor_name", "")],
        ["Quotation Ref:", po_data.get("quotation_no", "")],
    ]
    
    info_table = Table(po_info, colWidths=[120, 300])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Items table
    items = po_data.get("items", [])
    if items:
        table_data = [["S.No", "Description", "Qty", "Unit Price", "Amount"]]
        for i, item in enumerate(items, 1):
            table_data.append([
                str(i),
                item.get("description", ""),
                str(item.get("quantity", 1)),
                f"₹{item.get('unit_price', 0):,.2f}",
                f"₹{item.get('amount', 0):,.2f}"
            ])
        
        items_table = Table(table_data, colWidths=[40, 250, 50, 80, 80])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(items_table)
    
    elements.append(Spacer(1, 20))
    
    # Totals
    subtotal = po_data.get("subtotal", po_data.get("total_amount", 0))
    gst = po_data.get("gst_amount", subtotal * po_data.get("gst_pct", 18) / 100)
    total = po_data.get("total_amount", subtotal + gst)
    
    totals_data = [
        ["Subtotal:", f"₹{subtotal:,.2f}"],
        [f"GST ({po_data.get('gst_pct', 18)}%):", f"₹{gst:,.2f}"],
        ["Total:", f"₹{total:,.2f}"],
    ]
    
    totals_table = Table(totals_data, colWidths=[380, 100])
    totals_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(totals_table)
    
    elements.append(Spacer(1, 30))
    
    # Terms
    elements.append(Paragraph("<b>Terms & Conditions:</b>", normal_style))
    elements.append(Paragraph(f"Payment Terms: {po_data.get('payment_terms', 'Net 30')}", normal_style))
    elements.append(Paragraph(f"Delivery Address: {po_data.get('delivery_address', 'NucleoVir Therapeutics Pvt. Ltd.')}", normal_style))
    
    elements.append(Spacer(1, 50))
    
    # Signatures
    sig_data = [["Prepared By", "Approved By"],
                ["_________________", "_________________"],
                ["CA/Finance", "Director"]]
    
    sig_table = Table(sig_data, colWidths=[240, 240])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 20),
    ]))
    elements.append(sig_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

# Database connection helper
def get_db():
    from server import db
    return db

def get_user_from_token_sync(session_token):
    from server import get_user_from_token
    return get_user_from_token

# ==================== QUOTATION ENDPOINTS ====================

@procurement_router.post("/quotations/upload")
async def upload_quotation(
    file: UploadFile = File(...),
    vendor_name: str = Form(None),
    category: str = Form("General"),
    department: str = Form("Operations"),
    session_token: Optional[str] = Cookie(None)
):
    """Upload quotation file with OCR extraction"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Check CA access
    if user.email.lower() != CA_EMAIL.lower() and user.role != "Admin":
        raise HTTPException(status_code=403, detail="Only CA can upload quotations")
    
    # Read file
    file_content = await file.read()
    
    # Perform OCR if PDF
    ocr_result = {}
    if file.filename.lower().endswith('.pdf'):
        ocr_result = ocr_extract_from_pdf(file_content)
    
    # Store file
    storage_path = get_storage_path("quotations")
    file_id = f"quo_{uuid.uuid4().hex[:12]}"
    file_ext = Path(file.filename).suffix
    stored_filename = f"{file_id}{file_ext}"
    file_path = storage_path / stored_filename
    
    with open(file_path, 'wb') as f:
        f.write(file_content)
    
    # Create quotation record
    quotation_doc = {
        "quotation_id": file_id,
        "quotation_no": ocr_result.get("fields", {}).get("quotation_no", {}).get("value", f"QUO-{uuid.uuid4().hex[:8].upper()}"),
        "vendor_name": vendor_name or ocr_result.get("fields", {}).get("vendor_name", {}).get("value", "Unknown Vendor"),
        "category": category,
        "gst_pct": ocr_result.get("fields", {}).get("gst_pct", {}).get("value", 18.0),
        "total_amount": ocr_result.get("fields", {}).get("total_amount", {}).get("value", 0),
        "validity_date": ocr_result.get("fields", {}).get("validity_date", {}).get("value", ""),
        "department": department,
        "description": "",
        "items": [],
        "file_path": str(file_path),
        "file_name": file.filename,
        "ocr_data": ocr_result,
        "ocr_confidence": ocr_result.get("confidence", 0),
        "status": "draft",
        "created_by": user.email,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "archived": False
    }
    
    await db.quotations.insert_one(quotation_doc)
    
    # Audit log
    await create_audit_log(db, "quotation", file_id, "created", user.email, None, {"status": "draft"})
    
    return {
        "quotation_id": file_id,
        "ocr_result": ocr_result,
        "needs_confirmation": ocr_result.get("confidence", 0) < 80,
        "extracted_fields": ocr_result.get("fields", {}),
        "message": "Quotation uploaded. Please confirm extracted data."
    }

@procurement_router.post("/quotations")
async def create_or_update_quotation(
    data: QuotationCreate,
    quotation_id: Optional[str] = None,
    session_token: Optional[str] = Cookie(None)
):
    """Create or update quotation with manual data"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    if user.email.lower() != CA_EMAIL.lower() and user.role != "Admin":
        raise HTTPException(status_code=403, detail="Only CA can manage quotations")
    
    if quotation_id:
        # Update existing
        old_doc = await db.quotations.find_one({"quotation_id": quotation_id}, {"_id": 0})
        if not old_doc:
            raise HTTPException(status_code=404, detail="Quotation not found")
        
        update_data = data.model_dump()
        update_data["status"] = "confirmed"
        update_data["confirmed_by"] = user.email
        update_data["confirmed_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.quotations.update_one(
            {"quotation_id": quotation_id},
            {"$set": update_data}
        )
        
        await create_audit_log(db, "quotation", quotation_id, "updated", user.email, old_doc, update_data)
        
        return {"message": "Quotation updated", "quotation_id": quotation_id}
    else:
        # Create new
        new_id = f"quo_{uuid.uuid4().hex[:12]}"
        quotation_doc = data.model_dump()
        quotation_doc["quotation_id"] = new_id
        quotation_doc["status"] = "confirmed"
        quotation_doc["created_by"] = user.email
        quotation_doc["created_at"] = datetime.now(timezone.utc).isoformat()
        quotation_doc["archived"] = False
        
        await db.quotations.insert_one(quotation_doc)
        await create_audit_log(db, "quotation", new_id, "created", user.email, None, {"status": "confirmed"})
        
        return {"message": "Quotation created", "quotation_id": new_id}

@procurement_router.get("/quotations")
async def list_quotations(
    status: Optional[str] = None,
    session_token: Optional[str] = Cookie(None)
):
    """List all quotations"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    query = {"archived": {"$ne": True}}
    if status:
        query["status"] = status
    
    quotations = await db.quotations.find(query, {"_id": 0}).sort([("created_at", -1)]).to_list(500)
    return quotations

@procurement_router.get("/quotations/{quotation_id}")
async def get_quotation(quotation_id: str, session_token: Optional[str] = Cookie(None)):
    """Get single quotation"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    quotation = await db.quotations.find_one({"quotation_id": quotation_id}, {"_id": 0})
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    return quotation

# ==================== PURCHASE ORDER ENDPOINTS ====================

@procurement_router.post("/po/generate")
async def generate_po(data: POGenerate, session_token: Optional[str] = Cookie(None)):
    """Generate PO from quotation"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    if user.email.lower() != CA_EMAIL.lower() and user.role != "Admin":
        raise HTTPException(status_code=403, detail="Only CA can generate POs")
    
    # Get quotation
    quotation = await db.quotations.find_one({"quotation_id": data.quotation_id}, {"_id": 0})
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    amount = quotation.get("total_amount", 0)
    
    # Check 3-quotes rule for ₹50,001-₹2,00,000
    if THRESHOLDS["single_director"] < amount <= THRESHOLDS["single_director_3quotes"]:
        # Check if there are 3 quotations for same vendor/category
        related_quotations = await db.quotations.count_documents({
            "vendor_name": quotation["vendor_name"],
            "category": quotation["category"],
            "status": "confirmed",
            "archived": {"$ne": True}
        })
        
        if related_quotations < 3:
            raise HTTPException(
                status_code=400,
                detail=f"3 quotations required for amounts ₹50,001-₹2,00,000. Currently have {related_quotations} quotations."
            )
    
    # Check board resolution for > ₹10,00,000
    if amount > THRESHOLDS["all_directors"]:
        # Board resolution will be required during approval
        pass
    
    # Get next PO number
    counter = await db.counters.find_one_and_update(
        {"_id": "po_counter"},
        {"$inc": {"value": 1}},
        upsert=True,
        return_document=True
    )
    po_number = generate_po_number(counter.get("value", 1))
    
    # Create PO
    po_id = f"po_{uuid.uuid4().hex[:12]}"
    po_doc = {
        "po_id": po_id,
        "po_number": po_number,
        "quotation_id": data.quotation_id,
        "quotation_no": quotation.get("quotation_no"),
        "vendor_name": quotation.get("vendor_name"),
        "category": quotation.get("category"),
        "department": quotation.get("department"),
        "items": quotation.get("items", []),
        "subtotal": amount / (1 + quotation.get("gst_pct", 18) / 100),
        "gst_pct": quotation.get("gst_pct", 18),
        "gst_amount": amount - (amount / (1 + quotation.get("gst_pct", 18) / 100)),
        "total_amount": amount,
        "delivery_address": data.delivery_address,
        "payment_terms": data.payment_terms,
        "notes": data.notes,
        "status": "pending_approval",
        "approval_status": "pending",
        "created_by": user.email,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "archived": False
    }
    
    await db.purchase_orders.insert_one(po_doc)
    
    # Update quotation status
    await db.quotations.update_one(
        {"quotation_id": data.quotation_id},
        {"$set": {"status": "converted_to_po", "po_id": po_id}}
    )
    
    # Create approval entries
    await create_approval_entries(db, po_id, amount, user.email)
    
    # Audit log
    await create_audit_log(db, "purchase_order", po_id, "created", user.email, None, {"status": "pending_approval"})
    
    return {
        "message": "Purchase Order generated",
        "po_id": po_id,
        "po_number": po_number,
        "approval_required": True
    }

@procurement_router.get("/po")
async def list_purchase_orders(
    status: Optional[str] = None,
    session_token: Optional[str] = Cookie(None)
):
    """List purchase orders"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    query = {"archived": {"$ne": True}}
    if status:
        query["status"] = status
    
    pos = await db.purchase_orders.find(query, {"_id": 0}).sort([("created_at", -1)]).to_list(500)
    return pos

@procurement_router.get("/po/{po_id}")
async def get_purchase_order(po_id: str, session_token: Optional[str] = Cookie(None)):
    """Get single PO"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    po = await db.purchase_orders.find_one({"po_id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="PO not found")
    
    # Get approvals
    approvals = await db.approvals.find({"po_id": po_id}, {"_id": 0}).sort([("approval_order", 1)]).to_list(10)
    po["approvals"] = approvals
    
    return po

@procurement_router.get("/po/{po_id}/pdf")
async def get_po_pdf(po_id: str, session_token: Optional[str] = Cookie(None)):
    """Generate and download PO PDF"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    po = await db.purchase_orders.find_one({"po_id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="PO not found")
    
    pdf_bytes = generate_po_pdf(po)
    if not pdf_bytes:
        raise HTTPException(status_code=500, detail="PDF generation not available")
    
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={po['po_number'].replace('/', '_')}.pdf"}
    )

# ==================== APPROVAL ENDPOINTS ====================

@procurement_router.get("/approvals/pending")
async def get_pending_approvals(
    approver_email: Optional[str] = None,
    session_token: Optional[str] = Cookie(None)
):
    """Get pending approvals for a user"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    email = approver_email or user.email
    
    # Get pending approvals for this user
    approvals = await db.approvals.find({
        "approver_email": email.lower(),
        "status": "pending"
    }, {"_id": 0}).to_list(100)
    
    # Enrich with PO data
    result = []
    for approval in approvals:
        po = await db.purchase_orders.find_one({"po_id": approval["po_id"]}, {"_id": 0})
        if po:
            approval["po_details"] = po
            result.append(approval)
    
    return result

@procurement_router.post("/approvals/{approval_id}/decision")
async def make_approval_decision(
    approval_id: str,
    data: ApprovalDecision,
    session_token: Optional[str] = Cookie(None)
):
    """Record approval decision"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Verify approver
    approval = await db.approvals.find_one({"approval_id": approval_id}, {"_id": 0})
    if not approval:
        raise HTTPException(status_code=404, detail="Approval not found")
    
    if approval["approver_email"].lower() != user.email.lower():
        raise HTTPException(status_code=403, detail="You are not authorized to approve this")
    
    if approval["status"] != "pending":
        raise HTTPException(status_code=400, detail="This approval is not pending")
    
    # Update approval
    old_status = approval["status"]
    await db.approvals.update_one(
        {"approval_id": approval_id},
        {"$set": {
            "status": data.decision,
            "decision": data.decision,
            "comment": data.comment,
            "decided_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Audit log
    await create_audit_log(
        db, "approval", approval_id, f"decision_{data.decision}", 
        user.email, 
        {"status": old_status}, 
        {"status": data.decision, "comment": data.comment}
    )
    
    po_id = approval["po_id"]
    po = await db.purchase_orders.find_one({"po_id": po_id}, {"_id": 0})
    
    if data.decision == "approved":
        # Check if there are more approvers
        next_approval = await db.approvals.find_one({
            "po_id": po_id,
            "status": "waiting",
            "approval_order": approval["approval_order"] + 1
        })
        
        if next_approval:
            # Activate next approver
            await db.approvals.update_one(
                {"approval_id": next_approval["approval_id"]},
                {"$set": {"status": "pending"}}
            )
            
            # Notify next approver
            await create_notification(
                db,
                next_approval["approver_email"],
                "PO Pending Your Approval",
                f"PO {po['po_number']} for ₹{po['total_amount']:,.2f} requires your approval.",
                "approval_pending",
                {"type": "po", "id": po_id}
            )
        else:
            # All approved - update PO status
            await db.purchase_orders.update_one(
                {"po_id": po_id},
                {"$set": {"status": "approved", "approval_status": "approved", "approved_at": datetime.now(timezone.utc).isoformat()}}
            )
            
            # Notify CA
            await create_notification(
                db,
                CA_EMAIL,
                "PO Approved",
                f"PO {po['po_number']} has been fully approved.",
                "po_approved",
                {"type": "po", "id": po_id}
            )
            
            await create_audit_log(db, "purchase_order", po_id, "approved", user.email, {"status": "pending_approval"}, {"status": "approved"})
    
    elif data.decision == "rejected":
        # Reject entire PO
        await db.purchase_orders.update_one(
            {"po_id": po_id},
            {"$set": {"status": "rejected", "approval_status": "rejected", "rejected_at": datetime.now(timezone.utc).isoformat(), "rejection_reason": data.comment}}
        )
        
        # Archive (not delete)
        await db.purchase_orders.update_one(
            {"po_id": po_id},
            {"$set": {"archived": True, "archive_reason": f"Rejected by {user.email}: {data.comment}"}}
        )
        
        # Notify CA
        await create_notification(
            db,
            CA_EMAIL,
            "PO Rejected",
            f"PO {po['po_number']} was rejected by {user.name or user.email}. Reason: {data.comment}",
            "po_rejected",
            {"type": "po", "id": po_id}
        )
        
        await create_audit_log(db, "purchase_order", po_id, "rejected", user.email, {"status": "pending_approval"}, {"status": "rejected"})
    
    return {"message": f"Approval {data.decision}", "po_id": po_id}

# ==================== GRN ENDPOINTS ====================

@procurement_router.post("/grn")
async def create_grn(data: GRNCreate, session_token: Optional[str] = Cookie(None)):
    """Create Goods Receipt Note"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    if user.email.lower() != CA_EMAIL.lower() and user.role != "Admin":
        raise HTTPException(status_code=403, detail="Only CA can create GRN")
    
    # Verify PO exists and is approved
    po = await db.purchase_orders.find_one({"po_id": data.po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="PO not found")
    
    if po["status"] != "approved":
        raise HTTPException(status_code=400, detail="PO must be approved before creating GRN")
    
    # Get GRN number
    counter = await db.counters.find_one_and_update(
        {"_id": "grn_counter"},
        {"$inc": {"value": 1}},
        upsert=True,
        return_document=True
    )
    grn_number = generate_grn_number(counter.get("value", 1))
    
    grn_id = f"grn_{uuid.uuid4().hex[:12]}"
    grn_doc = {
        "grn_id": grn_id,
        "grn_number": grn_number,
        "po_id": data.po_id,
        "po_number": po["po_number"],
        "vendor_name": po["vendor_name"],
        "received_items": data.received_items,
        "qc_status": data.qc_status,
        "qc_notes": data.qc_notes,
        "asset_tags": data.asset_tags,
        "received_by": user.email,
        "received_at": datetime.now(timezone.utc).isoformat(),
        "status": "received",
        "archived": False
    }
    
    await db.grn.insert_one(grn_doc)
    
    # Update PO status
    await db.purchase_orders.update_one(
        {"po_id": data.po_id},
        {"$set": {"status": "goods_received", "grn_id": grn_id}}
    )
    
    # Create asset records if tags provided
    for i, tag in enumerate(data.asset_tags or []):
        asset_doc = {
            "asset_id": f"asset_{uuid.uuid4().hex[:12]}",
            "asset_tag": tag,
            "grn_id": grn_id,
            "po_id": data.po_id,
            "description": data.received_items[i].get("description", "") if i < len(data.received_items) else "",
            "status": "active",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.assets.insert_one(asset_doc)
    
    await create_audit_log(db, "grn", grn_id, "created", user.email, None, {"status": "received"})
    
    return {"message": "GRN created", "grn_id": grn_id, "grn_number": grn_number}

@procurement_router.get("/grn")
async def list_grns(session_token: Optional[str] = Cookie(None)):
    """List all GRNs"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    grns = await db.grn.find({"archived": {"$ne": True}}, {"_id": 0}).sort([("received_at", -1)]).to_list(500)
    return grns

# ==================== VOUCHER ENDPOINTS ====================

@procurement_router.post("/vouchers/create")
async def create_voucher(data: VoucherCreate, session_token: Optional[str] = Cookie(None)):
    """Create payment voucher after invoice + GRN"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    if user.email.lower() != CA_EMAIL.lower() and user.role != "Admin":
        raise HTTPException(status_code=403, detail="Only CA can create vouchers")
    
    # Verify GRN
    grn = await db.grn.find_one({"grn_id": data.grn_id}, {"_id": 0})
    if not grn:
        raise HTTPException(status_code=404, detail="GRN not found")
    
    # Get PO
    po = await db.purchase_orders.find_one({"po_id": grn["po_id"]}, {"_id": 0})
    
    # Get voucher number
    counter = await db.counters.find_one_and_update(
        {"_id": "voucher_counter"},
        {"$inc": {"value": 1}},
        upsert=True,
        return_document=True
    )
    voucher_number = generate_voucher_number(counter.get("value", 1))
    
    # Calculate TDS
    tds_amount = data.invoice_amount * (data.tds_pct or 0) / 100
    net_payable = data.invoice_amount - tds_amount
    
    voucher_id = f"vch_{uuid.uuid4().hex[:12]}"
    voucher_doc = {
        "voucher_id": voucher_id,
        "voucher_number": voucher_number,
        "grn_id": data.grn_id,
        "po_id": grn["po_id"],
        "po_number": po["po_number"],
        "vendor_name": po["vendor_name"],
        "invoice_number": data.invoice_number,
        "invoice_date": data.invoice_date,
        "invoice_amount": data.invoice_amount,
        "tds_pct": data.tds_pct or 0,
        "tds_amount": tds_amount,
        "net_payable": net_payable,
        "payment_mode": data.payment_mode,
        "status": "pending_review",
        "payment_status": "unpaid",
        "created_by": user.email,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "archived": False
    }
    
    await db.vouchers.insert_one(voucher_doc)
    
    await create_audit_log(db, "voucher", voucher_id, "created", user.email, None, {"status": "pending_review"})
    
    return {"message": "Voucher created", "voucher_id": voucher_id, "voucher_number": voucher_number}

@procurement_router.get("/vouchers")
async def list_vouchers(
    status: Optional[str] = None,
    session_token: Optional[str] = Cookie(None)
):
    """List vouchers"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    query = {"archived": {"$ne": True}}
    if status:
        query["status"] = status
    
    vouchers = await db.vouchers.find(query, {"_id": 0}).sort([("created_at", -1)]).to_list(500)
    return vouchers

@procurement_router.post("/vouchers/{voucher_id}/approve")
async def approve_voucher(voucher_id: str, session_token: Optional[str] = Cookie(None)):
    """CA approves voucher for payment"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    if user.email.lower() != CA_EMAIL.lower() and user.role != "Admin":
        raise HTTPException(status_code=403, detail="Only CA can approve vouchers")
    
    voucher = await db.vouchers.find_one({"voucher_id": voucher_id}, {"_id": 0})
    if not voucher:
        raise HTTPException(status_code=404, detail="Voucher not found")
    
    await db.vouchers.update_one(
        {"voucher_id": voucher_id},
        {"$set": {
            "status": "approved",
            "approved_by": user.email,
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await create_audit_log(db, "voucher", voucher_id, "approved", user.email, {"status": "pending_review"}, {"status": "approved"})
    
    return {"message": "Voucher approved"}

@procurement_router.post("/vouchers/{voucher_id}/mark-paid")
async def mark_voucher_paid(
    voucher_id: str,
    payment_reference: str = Form(...),
    payment_date: str = Form(...),
    session_token: Optional[str] = Cookie(None)
):
    """Mark voucher as paid"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    if user.email.lower() != CA_EMAIL.lower() and user.role != "Admin":
        raise HTTPException(status_code=403, detail="Only CA can mark payments")
    
    await db.vouchers.update_one(
        {"voucher_id": voucher_id},
        {"$set": {
            "payment_status": "paid",
            "payment_reference": payment_reference,
            "payment_date": payment_date,
            "paid_by": user.email,
            "paid_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await create_audit_log(db, "voucher", voucher_id, "paid", user.email, {"payment_status": "unpaid"}, {"payment_status": "paid"})
    
    return {"message": "Payment recorded"}

# ==================== REPORTS ENDPOINTS ====================

@procurement_router.get("/reports/export")
async def export_report(
    report_type: str = Query(..., description="po_register, payment_register, vendor_aging, gst_tds"),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    format: str = Query("csv", description="csv, xlsx, pdf"),
    session_token: Optional[str] = Cookie(None)
):
    """Export reports - CA and Admin only"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Block Directors from accessing reports (CA-only feature)
    if user.email.lower() != CA_EMAIL.lower() and user.role != "Admin":
        raise HTTPException(status_code=403, detail="Only CA and Admin can access reports")
    
    # Build date query
    date_query = {}
    if from_date:
        date_query["$gte"] = from_date
    if to_date:
        date_query["$lte"] = to_date
    
    data = []
    headers = []
    filename = f"{report_type}_{datetime.now().strftime('%Y%m%d')}"
    
    if report_type == "po_register":
        query = {"archived": {"$ne": True}}
        if date_query:
            query["created_at"] = date_query
        
        pos = await db.purchase_orders.find(query, {"_id": 0}).sort([("created_at", -1)]).to_list(1000)
        headers = ["PO Number", "Date", "Vendor", "Category", "Amount", "GST", "Total", "Status"]
        data = [[
            po.get("po_number", ""),
            po.get("created_at", "")[:10],
            po.get("vendor_name", ""),
            po.get("category", ""),
            po.get("subtotal", 0),
            po.get("gst_amount", 0),
            po.get("total_amount", 0),
            po.get("status", "")
        ] for po in pos]
    
    elif report_type == "payment_register":
        query = {"payment_status": "paid", "archived": {"$ne": True}}
        if date_query:
            query["paid_at"] = date_query
        
        vouchers = await db.vouchers.find(query, {"_id": 0}).sort([("paid_at", -1)]).to_list(1000)
        headers = ["Voucher No", "Date", "Vendor", "Invoice No", "Amount", "TDS", "Net Paid", "Payment Ref"]
        data = [[
            v.get("voucher_number", ""),
            v.get("payment_date", ""),
            v.get("vendor_name", ""),
            v.get("invoice_number", ""),
            v.get("invoice_amount", 0),
            v.get("tds_amount", 0),
            v.get("net_payable", 0),
            v.get("payment_reference", "")
        ] for v in vouchers]
    
    elif report_type == "vendor_aging":
        # Get unpaid vouchers grouped by vendor
        pipeline = [
            {"$match": {"payment_status": "unpaid", "archived": {"$ne": True}}},
            {"$group": {
                "_id": "$vendor_name",
                "total_due": {"$sum": "$net_payable"},
                "count": {"$sum": 1},
                "oldest_date": {"$min": "$created_at"}
            }}
        ]
        aging = await db.vouchers.aggregate(pipeline).to_list(100)
        headers = ["Vendor", "Total Due", "Invoices", "Oldest Invoice"]
        data = [[
            a.get("_id", ""),
            a.get("total_due", 0),
            a.get("count", 0),
            a.get("oldest_date", "")[:10] if a.get("oldest_date") else ""
        ] for a in aging]
    
    elif report_type == "gst_tds":
        query = {"archived": {"$ne": True}}
        if date_query:
            query["created_at"] = date_query
        
        vouchers = await db.vouchers.find(query, {"_id": 0}).to_list(1000)
        headers = ["Voucher No", "Vendor", "Invoice Amount", "GST Amount", "TDS %", "TDS Amount"]
        
        for v in vouchers:
            # Get PO for GST info
            po = await db.purchase_orders.find_one({"po_id": v.get("po_id")}, {"_id": 0})
            gst_amount = po.get("gst_amount", 0) if po else 0
            data.append([
                v.get("voucher_number", ""),
                v.get("vendor_name", ""),
                v.get("invoice_amount", 0),
                gst_amount,
                v.get("tds_pct", 0),
                v.get("tds_amount", 0)
            ])
    
    # Generate output
    if format == "csv":
        output = io.StringIO()
        output.write(",".join(headers) + "\n")
        for row in data:
            output.write(",".join(str(cell) for cell in row) + "\n")
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode()),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
        )
    
    elif format == "xlsx" and XLSX_AVAILABLE:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type
        
        ws.append(headers)
        for row in data:
            ws.append(row)
        
        # Auto-width columns
        for i, col in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )
    
    elif format == "pdf" and PDF_AVAILABLE:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"<b>{report_type.replace('_', ' ').title()}</b>", styles['Heading1']))
        elements.append(Spacer(1, 20))
        
        table_data = [headers] + data
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"}
        )
    
    return {"headers": headers, "data": data}

# ==================== COMBINED REPORT EXPORT ====================

@procurement_router.get("/reports/combined")
async def export_combined_report(
    from_date: str = Query(..., description="Start date YYYY-MM-DD"),
    to_date: str = Query(..., description="End date YYYY-MM-DD"),
    format: str = Query("csv", description="csv, xlsx, pdf"),
    session_token: Optional[str] = Cookie(None)
):
    """Export all reports combined into one file - CA and Admin only"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Block Directors from accessing reports (CA-only feature)
    if user.email.lower() != CA_EMAIL.lower() and user.role != "Admin":
        raise HTTPException(status_code=403, detail="Only CA and Admin can access reports")
    
    COMPANY_NAME = "NucleoVir Therapeutics Pvt. Ltd."
    date_range_str = f"{from_date} to {to_date}"
    filename = f"Reports_PO_Payment_Vendor_GSTTDS_{from_date}_{to_date}"
    
    # Ensure storage directory exists
    storage_dir = Path("/app/storage/Finance/Reports/FY2026")
    storage_dir.mkdir(parents=True, exist_ok=True)
    
    # Build date query
    date_query = {"$gte": from_date, "$lte": to_date}
    
    # ========== PO REGISTER ==========
    po_query = {"archived": {"$ne": True}, "created_at": date_query}
    pos = await db.purchase_orders.find(po_query, {"_id": 0}).sort([("created_at", -1)]).to_list(1000)
    
    po_headers = ["PO Number", "Date", "Vendor", "Category", "Amount", "GST", "Total", "Status"]
    po_data = [[
        po.get("po_number", ""),
        po.get("created_at", "")[:10] if po.get("created_at") else "",
        po.get("vendor_name", ""),
        po.get("category", ""),
        po.get("subtotal", 0),
        po.get("gst_amount", 0),
        po.get("total_amount", 0),
        po.get("status", "")
    ] for po in pos]
    
    # PO Totals
    po_total_amount = sum(p.get("subtotal", 0) for p in pos)
    po_total_gst = sum(p.get("gst_amount", 0) for p in pos)
    po_total_grand = sum(p.get("total_amount", 0) for p in pos)
    po_totals = ["TOTAL", "", "", "", po_total_amount, po_total_gst, po_total_grand, ""]
    
    # ========== PAYMENT REGISTER ==========
    payment_query = {"archived": {"$ne": True}, "payment_date": date_query}
    vouchers_paid = await db.vouchers.find(payment_query, {"_id": 0}).sort([("payment_date", -1)]).to_list(1000)
    
    payment_headers = ["Voucher No", "Date", "Vendor", "Invoice No", "Amount", "TDS", "Net Paid", "Payment Ref"]
    payment_data = [[
        v.get("voucher_number", ""),
        v.get("payment_date", ""),
        v.get("vendor_name", ""),
        v.get("invoice_number", ""),
        v.get("invoice_amount", 0),
        v.get("tds_amount", 0),
        v.get("net_payable", 0),
        v.get("payment_reference", "")
    ] for v in vouchers_paid]
    
    # Payment Totals
    pay_total_amount = sum(v.get("invoice_amount", 0) for v in vouchers_paid)
    pay_total_tds = sum(v.get("tds_amount", 0) for v in vouchers_paid)
    pay_total_net = sum(v.get("net_payable", 0) for v in vouchers_paid)
    payment_totals = ["TOTAL", "", "", "", pay_total_amount, pay_total_tds, pay_total_net, ""]
    
    # ========== VENDOR AGING ==========
    pipeline = [
        {"$match": {"payment_status": "unpaid", "archived": {"$ne": True}}},
        {"$group": {
            "_id": "$vendor_name",
            "total_due": {"$sum": "$net_payable"},
            "count": {"$sum": 1},
            "oldest_date": {"$min": "$created_at"}
        }}
    ]
    aging = await db.vouchers.aggregate(pipeline).to_list(100)
    
    aging_headers = ["Vendor", "Total Due", "Invoices", "Oldest Invoice"]
    aging_data = [[
        a.get("_id", ""),
        a.get("total_due", 0),
        a.get("count", 0),
        a.get("oldest_date", "")[:10] if a.get("oldest_date") else ""
    ] for a in aging]
    
    # Aging Totals
    aging_total_due = sum(a.get("total_due", 0) for a in aging)
    aging_total_invoices = sum(a.get("count", 0) for a in aging)
    aging_totals = ["TOTAL", aging_total_due, aging_total_invoices, ""]
    
    # ========== GST/TDS SUMMARY ==========
    gst_query = {"archived": {"$ne": True}, "created_at": date_query}
    all_vouchers = await db.vouchers.find(gst_query, {"_id": 0}).to_list(1000)
    
    gst_headers = ["Voucher No", "Vendor", "Invoice Amount", "GST Amount", "TDS %", "TDS Amount"]
    gst_data = []
    for v in all_vouchers:
        po = await db.purchase_orders.find_one({"po_id": v.get("po_id")}, {"_id": 0})
        gst_amount = po.get("gst_amount", 0) if po else 0
        gst_data.append([
            v.get("voucher_number", ""),
            v.get("vendor_name", ""),
            v.get("invoice_amount", 0),
            gst_amount,
            v.get("tds_pct", 0),
            v.get("tds_amount", 0)
        ])
    
    # GST/TDS Totals
    gst_total_invoice = sum(v.get("invoice_amount", 0) for v in all_vouchers)
    gst_total_gst = sum(row[3] for row in gst_data)
    gst_total_tds = sum(v.get("tds_amount", 0) for v in all_vouchers)
    gst_totals = ["TOTAL", "", gst_total_invoice, gst_total_gst, "", gst_total_tds]
    
    # ========== GENERATE OUTPUT ==========
    if format == "csv":
        output = io.StringIO()
        
        # Header
        output.write(f"{COMPANY_NAME}\n")
        output.write(f"Combined Financial Report\n")
        output.write(f"Period: {date_range_str}\n")
        output.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n")
        
        # PO Register
        output.write("=== PO REGISTER ===\n")
        output.write(",".join(po_headers) + "\n")
        for row in po_data:
            output.write(",".join(str(cell) for cell in row) + "\n")
        output.write(",".join(str(cell) for cell in po_totals) + "\n\n")
        
        # Payment Register
        output.write("=== PAYMENT REGISTER ===\n")
        output.write(",".join(payment_headers) + "\n")
        for row in payment_data:
            output.write(",".join(str(cell) for cell in row) + "\n")
        output.write(",".join(str(cell) for cell in payment_totals) + "\n\n")
        
        # Vendor Aging
        output.write("=== VENDOR AGING ===\n")
        output.write(",".join(aging_headers) + "\n")
        for row in aging_data:
            output.write(",".join(str(cell) for cell in row) + "\n")
        output.write(",".join(str(cell) for cell in aging_totals) + "\n\n")
        
        # GST/TDS
        output.write("=== GST/TDS SUMMARY ===\n")
        output.write(",".join(gst_headers) + "\n")
        for row in gst_data:
            output.write(",".join(str(cell) for cell in row) + "\n")
        output.write(",".join(str(cell) for cell in gst_totals) + "\n")
        
        # Save to storage
        file_path = storage_dir / f"{filename}.csv"
        with open(file_path, 'w') as f:
            f.write(output.getvalue())
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode()),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
        )
    
    elif format == "xlsx" and XLSX_AVAILABLE:
        wb = openpyxl.Workbook()
        
        # PO Register Sheet
        ws_po = wb.active
        ws_po.title = "PO Register"
        ws_po.append([COMPANY_NAME])
        ws_po.append(["Combined Report - PO Register"])
        ws_po.append([f"Period: {date_range_str}"])
        ws_po.append([])
        ws_po.append(po_headers)
        for row in po_data:
            ws_po.append(row)
        ws_po.append(po_totals)
        for i in range(1, len(po_headers) + 1):
            ws_po.column_dimensions[get_column_letter(i)].width = 15
        
        # Payment Register Sheet
        ws_pay = wb.create_sheet("Payment Register")
        ws_pay.append([COMPANY_NAME])
        ws_pay.append(["Combined Report - Payment Register"])
        ws_pay.append([f"Period: {date_range_str}"])
        ws_pay.append([])
        ws_pay.append(payment_headers)
        for row in payment_data:
            ws_pay.append(row)
        ws_pay.append(payment_totals)
        for i in range(1, len(payment_headers) + 1):
            ws_pay.column_dimensions[get_column_letter(i)].width = 15
        
        # Vendor Aging Sheet
        ws_aging = wb.create_sheet("Vendor Aging")
        ws_aging.append([COMPANY_NAME])
        ws_aging.append(["Combined Report - Vendor Aging"])
        ws_aging.append([f"Period: {date_range_str}"])
        ws_aging.append([])
        ws_aging.append(aging_headers)
        for row in aging_data:
            ws_aging.append(row)
        ws_aging.append(aging_totals)
        for i in range(1, len(aging_headers) + 1):
            ws_aging.column_dimensions[get_column_letter(i)].width = 15
        
        # GST/TDS Sheet
        ws_gst = wb.create_sheet("GST-TDS Summary")
        ws_gst.append([COMPANY_NAME])
        ws_gst.append(["Combined Report - GST/TDS Summary"])
        ws_gst.append([f"Period: {date_range_str}"])
        ws_gst.append([])
        ws_gst.append(gst_headers)
        for row in gst_data:
            ws_gst.append(row)
        ws_gst.append(gst_totals)
        for i in range(1, len(gst_headers) + 1):
            ws_gst.column_dimensions[get_column_letter(i)].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Save to storage
        file_path = storage_dir / f"{filename}.xlsx"
        with open(file_path, 'wb') as f:
            output.seek(0)
            f.write(output.read())
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )
    
    elif format == "pdf" and PDF_AVAILABLE:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, alignment=1)
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=1)
        section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12, spaceAfter=10)
        
        # Header
        elements.append(Paragraph(COMPANY_NAME, title_style))
        elements.append(Paragraph("Combined Financial Report", subtitle_style))
        elements.append(Paragraph(f"Period: {date_range_str}", subtitle_style))
        elements.append(Spacer(1, 20))
        
        def make_table(headers, data, totals):
            table_data = [headers] + data + [totals]
            t = Table(table_data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.13, 0.37, 0.6)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            return t
        
        # PO Register
        elements.append(Paragraph("PO Register", section_style))
        if po_data:
            elements.append(make_table(po_headers, po_data, po_totals))
        else:
            elements.append(Paragraph("No records found", styles['Normal']))
        elements.append(Spacer(1, 15))
        
        # Payment Register
        elements.append(Paragraph("Payment Register", section_style))
        if payment_data:
            elements.append(make_table(payment_headers, payment_data, payment_totals))
        else:
            elements.append(Paragraph("No records found", styles['Normal']))
        elements.append(Spacer(1, 15))
        
        # Vendor Aging
        elements.append(Paragraph("Vendor Aging", section_style))
        if aging_data:
            elements.append(make_table(aging_headers, aging_data, aging_totals))
        else:
            elements.append(Paragraph("No records found", styles['Normal']))
        elements.append(Spacer(1, 15))
        
        # GST/TDS Summary
        elements.append(Paragraph("GST/TDS Summary", section_style))
        if gst_data:
            elements.append(make_table(gst_headers, gst_data, gst_totals))
        else:
            elements.append(Paragraph("No records found", styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        
        # Save to storage
        file_path = storage_dir / f"{filename}.pdf"
        with open(file_path, 'wb') as f:
            buffer.seek(0)
            f.write(buffer.read())
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"}
        )
    
    return {"error": "Format not supported"}

# ==================== AUDIT & NOTIFICATIONS ====================

@procurement_router.get("/audit/{entity_type}/{entity_id}")
async def get_audit_trail(
    entity_type: str,
    entity_id: str,
    session_token: Optional[str] = Cookie(None)
):
    """Get audit trail for an entity"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    logs = await db.audit_logs.find({
        "entity_type": entity_type,
        "entity_id": entity_id
    }, {"_id": 0}).sort([("timestamp", -1)]).to_list(100)
    
    return logs

@procurement_router.get("/notifications")
async def get_notifications(
    user_email: Optional[str] = None,
    unread_only: bool = False,
    session_token: Optional[str] = Cookie(None)
):
    """Get in-app notifications"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    email = user_email or user.email
    
    query = {"user_email": email.lower()}
    if unread_only:
        query["read"] = False
    
    notifications = await db.notifications.find(query, {"_id": 0}).sort([("created_at", -1)]).to_list(100)
    return notifications

@procurement_router.post("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, session_token: Optional[str] = Cookie(None)):
    """Mark notification as read"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    await db.notifications.update_one(
        {"notification_id": notification_id},
        {"$set": {"read": True}}
    )
    
    return {"message": "Notification marked as read"}

# ==================== CA ACCESS CHECK ====================

@procurement_router.get("/procurement/access")
async def check_procurement_access(session_token: Optional[str] = Cookie(None)):
    """Check if user has CA or Director access"""
    from server import db, get_user_from_token
    
    user = await get_user_from_token(session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    email = user.email.lower()
    
    is_ca = email == CA_EMAIL.lower()
    is_director = email in [d.lower() for d in DIRECTORS]
    is_admin = user.role == "Admin"
    
    # Directors who are NOT also Admin should only have Director access
    # CA has full procurement access
    # Admin has full access
    can_access_ca_features = is_ca or is_admin
    
    # Get pending approvals count for directors
    pending_count = 0
    if is_director or is_admin:
        pending_count = await db.approvals.count_documents({
            "approver_email": email,
            "status": "pending"
        })
    
    return {
        "is_ca": is_ca,
        "is_director": is_director,
        "is_admin": is_admin,
        "can_access_procurement": is_ca or is_admin,
        "can_access_ca_features": can_access_ca_features,  # Full CA features
        "can_approve": is_director or is_admin,  # Director approval access only
        "pending_approvals": pending_count,
        "email": email,
        "role": user.role
    }
